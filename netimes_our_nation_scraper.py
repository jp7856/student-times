"""
NE Times KIDS - Our Nation 섹션 한국 관련 기사 수집기
------------------------------------------------------
실행 방법:
  1. pip install selenium webdriver-manager openpyxl
  2. python netimes_our_nation_scraper.py
  3. 크롬 열리면 직접 로그인 → 터미널에서 Enter → 자동 수집!
"""

import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

MAIN_URL     = "https://www.netimes.co.kr/index.asp"
KIDS_URL     = "https://www.netimes.co.kr/pages/Kids/index.asp"
MAX_ARTICLES = 10
OUTPUT_FILE  = "netimes_our_nation.xlsx"


def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1280,900")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    return driver


def manual_login(driver):
    driver.get(MAIN_URL)
    print("\n🌐 크롬 브라우저가 열렸습니다!")
    print("👉 메인 페이지에서 직접 로그인해주세요.")
    input("✅ 로그인 완료 후 여기서 Enter 누르세요...")
    print(f"로그인 확인 (현재 URL: {driver.current_url})")


def get_article_links(driver):
    driver.get(KIDS_URL)
    time.sleep(3)
    links = []
    anchors = driver.find_elements(By.TAG_NAME, "a")
    for a in anchors:
        href = a.get_attribute("href") or ""
        text = a.text.strip().lower()
        parent_text = ""
        try:
            parent_text = a.find_element(By.XPATH, "./..").text.lower()
        except Exception:
            pass
        if "reading.asp" in href and ("our nation" in text or "our nation" in parent_text):
            if href not in links:
                links.append(href)
    print(f"📋 Our Nation 기사 링크 수집: {len(links)}개")
    return links


def search_our_nation_articles(driver):
    SEARCH_URL = "https://www.netimes.co.kr/pages/Kids/search.asp"
    results = []
    keywords = ["Korea", "Korean", "Seoul", "South Korea"]
    for kw in keywords:
        try:
            driver.get(f"{SEARCH_URL}?searchstr={kw}&searchtype=1")
            time.sleep(2)
            rows = driver.find_elements(By.CSS_SELECTOR, "table tr, .article_list li, .news_list li")
            for row in rows:
                text = row.text.strip()
                for a in row.find_elements(By.TAG_NAME, "a"):
                    href = a.get_attribute("href") or ""
                    if "reading.asp" in href and "seq=" in href:
                        title = a.text.strip() or text[:80]
                        if href not in [r["url"] for r in results]:
                            results.append({"title": title, "url": href})
        except Exception as e:
            print(f"  검색 오류 ({kw}): {e}")
    return results


def get_article_detail(driver, url):
    driver.get(url)
    time.sleep(2)
    detail = {"url": url, "title": "", "section": "Our Nation", "content": "", "date": ""}
    try:
        for sel in ["h1", "h2", ".article_title", ".news_title", ".title"]:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            if els and els[0].text.strip():
                detail["title"] = els[0].text.strip()
                break
        for sel in [".date", ".news_date", "span[class*='date']", "td.date"]:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            if els and els[0].text.strip():
                detail["date"] = els[0].text.strip()
                break
        for sel in [".section", ".category", ".sec_name", "span[class*='section']"]:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            if els and els[0].text.strip():
                detail["section"] = els[0].text.strip()
                break
        for sel in [".article_content", ".news_content", ".content", "#article_body", "div.text"]:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            if els and els[0].text.strip():
                detail["content"] = els[0].text.strip()[:500]
                break
        if not detail["content"]:
            body_text = driver.find_element(By.TAG_NAME, "body").text
            lines = [l.strip() for l in body_text.split("\n") if len(l.strip()) > 30]
            detail["content"] = " ".join(lines[2:8])[:500]
    except Exception as e:
        print(f"  상세 추출 오류: {e}")
    return detail


def save_to_excel(articles, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Our Nation 기사"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border_side  = Side(style="thin", color="CCCCCC")
    thin_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    headers    = ["#", "제목 (Title)", "섹션", "날짜", "본문 요약 (첫 500자)", "URL"]
    col_widths = [5, 40, 15, 12, 60, 50]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 25

    alt_fill = PatternFill("solid", start_color="EBF3FB")
    for row_idx, art in enumerate(articles, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        values = [
            row_idx - 1,
            art.get("title", ""),
            art.get("section", "Our Nation"),
            art.get("date", ""),
            art.get("content", ""),
            art.get("url", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center_align if col_idx in [1, 3, 4] else left_align
            cell.font = Font(name="Arial", bold=(col_idx == 1), size=10)
        ws.row_dimensions[row_idx].height = 80

    ws.freeze_panes = "A2"
    wb.save(filename)
    print(f"\n✅ 엑셀 저장 완료: {filename} ({len(articles)}개 기사)")


def main():
    print("=" * 50)
    print("  NE Times KIDS - Our Nation 기사 수집기")
    print("=" * 50)

    driver = init_driver()
    try:
        # 1. 수동 로그인
        manual_login(driver)

        # 2. 기사 링크 수집
        article_links = get_article_links(driver)
        search_results = search_our_nation_articles(driver)
        print(f"🔍 검색으로 추가 수집: {len(search_results)}개")

        all_urls = list(dict.fromkeys(
            article_links + [r["url"] for r in search_results]
        ))[:MAX_ARTICLES * 2]

        if not all_urls:
            print("❌ 기사를 찾지 못했습니다.")
            return

        # 3. 기사 상세 수집
        print(f"\n📰 기사 상세 수집 중... (최대 {MAX_ARTICLES}개)")
        articles = []
        for i, url in enumerate(all_urls):
            if len(articles) >= MAX_ARTICLES:
                break
            print(f"  [{i+1}] {url[:70]}...")
            detail = get_article_detail(driver, url)
            if detail["title"]:
                articles.append(detail)
                print(f"      → {detail['title'][:50]}")

        # 4. 엑셀 저장
        if articles:
            save_to_excel(articles, OUTPUT_FILE)
        else:
            print("❌ 수집된 기사가 없습니다.")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()