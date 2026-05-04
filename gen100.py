"""
NE Times KIDS - Our Nation 기사 100개 AI 생성기
"""

import anthropic
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SAMPLE = """
Our Nation / LEVEL 1 / Sky Bridge

A new bridge opened in Incheon. It is called the Cheongna Sky Bridge.
The bridge connects Yeongjong Island with Cheongna. It helps people travel faster and more safely.
The bridge also has a tall observatory. It is called the Sky 184 because it is 184 meters high.
Recently, the Sky 184 set a new Guinness World Record. It is the highest observatory on a sea bridge.

Q1: What is the name of the bridge?
Q2: What does the bridge connect?
Q3: Do you like visiting bridges or tall places?

Words: bridge, open, call, connect, travel, fast, safely, tall, observatory, recently, set, record
"""

TOPICS = [
    "K-pop, 한류, 음악, 드라마",
    "한국 음식, 요리, 식문화, 길거리 음식",
    "전통문화, 명절, 역사, 유적지",
    "스포츠, 태권도, 올림픽, e스포츠",
    "기술, 과학, 우주, 환경, 교육",
]

def generate_batch(client, topic_hint, count=20):
    prompt = f"""아래는 NE Times KIDS 영자신문 'Our Nation' 섹션 LEVEL 1 샘플 기사야.

[샘플]
{SAMPLE}

위 스타일과 동일하게, 한국(Korea) 관련 주제 중 [{topic_hint}] 분야에서
서로 다른 기사 {count}개를 써줘.

규칙:
- 섹션: Our Nation (고정)
- 레벨: LEVEL 1 (짧고 쉬운 영어 문장)
- 본문: 4~5문장, 2~3문단
- Discussion Questions: 3개
- Words: 핵심 단어 10개 (영어만, 콤마로 구분)
- 주제가 서로 겹치지 않게

반드시 아래 JSON 배열 형식으로만 응답해. 다른 말 하지 마:
[
  {{
    "title": "기사 제목",
    "content": "본문 (문단 구분은 \\n\\n 사용)",
    "q1": "Q1 질문",
    "q2": "Q2 질문",
    "q3": "Q3 질문",
    "words": "word1, word2, word3, ..."
  }}
]"""

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}]
    )

    raw = response.content[0].text.strip()
    try:
        return json.loads(raw)
    except Exception:
        match = re.search(r'\[.*\]', raw, re.DOTALL)
        if match:
            return json.loads(match.group())
        return []


def save_to_excel(articles, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Our Nation 기사"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border_side  = Side(style="thin", color="CCCCCC")
    thin_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    headers    = ["#", "섹션", "제목 (Title)", "본문 (Content)", "Q1", "Q2", "Q3", "단어 (Words)"]
    col_widths = [4, 12, 35, 55, 25, 25, 30, 50]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 25

    alt_fill = PatternFill("solid", start_color="EBF3FB")
    for row_idx, art in enumerate(articles, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        values = [
            row_idx - 1,
            "Our Nation",
            art.get("title", ""),
            art.get("content", ""),
            art.get("q1", ""),
            art.get("q2", ""),
            art.get("q3", ""),
            art.get("words", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center_align if col_idx in [1, 2] else left_align
            cell.font = Font(name="Arial", bold=(col_idx == 1), size=10)
        ws.row_dimensions[row_idx].height = 90

    ws.freeze_panes = "A2"
    wb.save(filename)
    print(f"\n✅ 엑셀 저장 완료: {filename} (총 {len(articles)}개)")


def main():
    print("=" * 55)
    print("  NE Times KIDS - Our Nation 기사 100개 생성기")
    print("=" * 55)

    api_key = input("\nClaude API 키 입력: ").strip()
    client = anthropic.Anthropic(api_key=api_key)
    all_articles = []

    for i, topic in enumerate(TOPICS):
        print(f"\n📝 배치 {i+1}/5 생성 중... ({topic})")
        batch = generate_batch(client, topic, count=20)
        all_articles.extend(batch)
        print(f"   ✅ {len(batch)}개 완료 (누적: {len(all_articles)}개)")

    save_to_excel(all_articles, "netimes_our_nation_100.xlsx")
    print("\n📂 같은 폴더에 netimes_our_nation_100.xlsx 저장됐어요!")


if __name__ == "__main__":
    main()